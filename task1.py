# coding: utf-8

import mysql.connector
import openpyxl
import datetime
import pandas as pd
import os


conn = mysql.connector.connect(host='localhost',database='my_db',user='my_db_user')


cur = conn.cursor()


APPLICATIONS_TABLE = 'applications'
CONTRACTS_TABLE = 'contracts'


cur.execute("""
CREATE TABLE IF NOT EXISTS %s (
    id_number INT AUTO_INCREMENT,
    name VARCHAR(100) NOT NULL,
    birth_date DATE NOT NULL,
    application_date DATE NOT NULL,
    gender BOOLEAN NOT NULL,
    employed_by VARCHAR(100),
    education VARCHAR(100),
    children INT,
    family INT,
    marital_status VARCHAR(20),
    city VARCHAR(50),
    position VARCHAR(100),
    income FLOAT,
    income_type VARCHAR(100),
    housing VARCHAR(100),
    house_ownership BOOLEAN,
    age_of_car INT,
    PRIMARY KEY (id_number)
)  ENGINE=INNODB;
""" % APPLICATIONS_TABLE)


cur.execute("""
CREATE TABLE IF NOT EXISTS %s (
    contract_number INT AUTO_INCREMENT,
    id_number INT NOT NULL,
    borrower VARCHAR(100) NOT NULL,
    contract_date DATE NOT NULL,
    amount FLOAT,
    type VARCHAR(100),
    term INT,
    annuity FLOAT,
    PRIMARY KEY (contract_number),
    FOREIGN KEY (id_number) REFERENCES %s(id_number) ON DELETE CASCADE
)  ENGINE=INNODB;
""" % (CONTRACTS_TABLE, APPLICATIONS_TABLE))


for file_name in os.listdir("анкеты"):
    full_path = os.path.join("анкеты", file_name)
    print("Loading file %s..." % full_path)
    if os.path.isfile(full_path):
        wb = openpyxl.load_workbook(full_path)
        sheet = wb[wb.sheetnames[0]]

        id_number = int(sheet.cell(6, 1).value)
        name = "'" + sheet.cell(4, 1).value + "'"

        birth_date = sheet.cell(4, 7).value
        if not(isinstance(birth_date, datetime.date) or isinstance(birth_date, datetime.datetime)):
            birth_date = datetime.datetime.strptime(birth_date, "%m.%d.%Y")
        else:
            birth_date = datetime.date(birth_date.year, birth_date.day, birth_date.month)

        application_date = sheet.cell(2, 15).value
        if not(isinstance(application_date, datetime.date) or isinstance(application_date, datetime.datetime)):
            application_date = datetime.datetime.strptime(application_date, "%m.%d.%Y")
        else:
            application_date = datetime.date(application_date.year, application_date.day, application_date.month)

        gender = sheet.cell(4, 8).value
        if gender not in ("Male", "Female"):
            print("Error in application %d: Wrong Gender!" % id_number)
            continue
        gender = True if gender == 'Female' else False

        employed_by = sheet.cell(4, 9).value
        employed_by = "'" + employed_by + "'" if employed_by else 'NULL'

        education = sheet.cell(6, 9).value
        education = "'" + education + "'" if education else 'NULL'

        children = sheet.cell(8, 5).value
        children = children if children else 'NULL'

        family = sheet.cell(8, 6).value
        family = family if family else 'NULL'

        marital_status = sheet.cell(8, 7).value
        marital_status = "'" + marital_status + "'" if marital_status else 'NULL'

        city = sheet.cell(8, 9).value
        city = "'" + city + "'" if city else 'NULL'

        position = sheet.cell(10, 13).value
        position = "'" + position + "'" if position else 'NULL'

        income = sheet.cell(15, 1).value
        income = income if income else 'NULL'

        income_type = sheet.cell(17, 1).value
        income_type = "'" + income_type + "'" if income_type else 'NULL'

        housing = sheet.cell(19, 1).value
        housing = "'" + housing + "'" if housing else 'NULL'

        house_ownership = sheet.cell(19, 5).value
        house_ownership = True if house_ownership == 'Y' else False

        age_of_car = sheet.cell(21, 1).value
        age_of_car = age_of_car if age_of_car else 'NULL'

        cur.execute("""
        INSERT INTO %s VALUES (%d, %s, '%s', '%s', %d, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %d, %s)
        """ % (APPLICATIONS_TABLE, id_number, name, birth_date, application_date, gender, employed_by, education, children, family, marital_status, city, position, income, income_type, housing, house_ownership, age_of_car))
conn.commit()


for file_name in os.listdir("контракты"):
    full_path = os.path.join("контракты", file_name)
    print("Loading file %s..." % full_path)
    if os.path.isfile(full_path):
        wb = openpyxl.load_workbook(full_path)
        sheet = wb[wb.sheetnames[0]]

        contract_number = int(sheet.cell(6, 5).value)
        id_number = int(sheet.cell(6, 1).value)
        borrower = sheet.cell(4, 1).value

        contract_date = sheet.cell(2, 6).value
        if not(isinstance(contract_date, datetime.date) or isinstance(contract_date, datetime.datetime)):
            contract_date = datetime.datetime.strptime(contract_date, "%m.%d.%Y")
        else:
            contract_date = datetime.date(contract_date.year, contract_date.day, contract_date.month)

        amount = sheet.cell(8, 1).value if sheet.cell(8, 1).value else 'NULL'
        type_c = sheet.cell(8, 5).value if sheet.cell(8, 5).value else 'NULL'
        term = sheet.cell(10, 1).value if sheet.cell(10, 1).value else 'NULL'
        annuity = sheet.cell(10, 5).value if sheet.cell(10, 5).value else 'NULL'

        try:
            cur.execute("""
            INSERT INTO %s VALUES (%d, %d, '%s', '%s', %s, '%s', %s, %s)
            """ % (CONTRACTS_TABLE, contract_number, id_number, borrower, contract_date, amount, type_c, term, annuity))
        except mysql.connector.IntegrityError:
            print("ERROR! There are no application with id=%d!" % id_number)
conn.commit()


print("-"*50)


print("Task 2.\nChecking values...")


print("Checking income...")
cur.execute("SELECT COUNT(*) FROM applications WHERE income < 0")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong incomes (<0)!")
    cur.execute("UPDATE applications SET income=-1*(income) WHERE income < 0")
    print("Updated applications with wrong income!")


print("Checking birth_date...")
cur.execute("SELECT COUNT(*) FROM applications WHERE birth_date < '1900-01-01' OR birth_date > CURDATE()")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong birth_date (<1900-01-01 OR > current date)!")
    cur.execute("DELETE FROM applications WHERE birth_date < '1900-01-01' OR birth_date > CURDATE()")
    print("Deleted applications with wrong birth_date!")


print("Checking children...")
cur.execute("SELECT COUNT(*) FROM applications WHERE children < 0")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong children count (<0)!")
    cur.execute("UPDATE applications SET children=-1*(children) WHERE children < 0")
    print("Updated applications with wrong children count!")


print("Checking family...")
cur.execute("SELECT COUNT(*) FROM applications WHERE family < 0")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong family count (<0)!")
    cur.execute("UPDATE applications SET family=-1*(family) WHERE family < 0")
    print("Updated applications with wrong family count!")


print("Checking age_of_car...")
cur.execute("SELECT COUNT(*) FROM applications WHERE age_of_car < 0")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong age_of_car (<0)!")
    cur.execute("UPDATE applications SET age_of_car=-1*(age_of_car) WHERE age_of_car < 0")
    print("Updated applications with wrong age_of_car!")


print("Checking children count for married people...")
cur.execute("SELECT COUNT(*) FROM applications WHERE marital_status IN ('Married', 'Civil marriage') AND children > 0 AND family <> (children + 2)")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong children count (children count + 2 is greater than family count)!")
    cur.execute("UPDATE applications SET family=children + 2 WHERE marital_status IN ('Married', 'Civil marriage') AND children > 0 AND family <> (children + 2)")
    print("Updated applications with children count!")


print("Checking amount...")
cur.execute("SELECT COUNT(*) FROM contracts WHERE amount < 0")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong amount (<0)!")
    cur.execute("UPDATE applications SET amount=-1*(amount) WHERE amount < 0")
    print("Updated contracts with wrong amount!")


print("Checking term...")
cur.execute("SELECT COUNT(*) FROM contracts WHERE term < 0")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong term (<0)!")
    cur.execute("UPDATE applications SET term=-1*(term) WHERE term < 0")
    print("Updated contracts with wrong term!")


print("Checking annuity...")
cur.execute("SELECT COUNT(*) FROM contracts WHERE annuity < 0")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong annuity (<0)!")
    cur.execute("UPDATE applications SET annuity=-1*(annuity) WHERE annuity < 0")
    print("Updated contracts with wrong annuity!")


print("Checking annuity is less than amount")
cur.execute("SELECT COUNT(*) FROM contracts WHERE annuity > amount")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong annuity (>amount)!")
    cur.execute("DELETE FROM contracts WHERE annuity > amount")
    print("Deleted contracts with wrong annuity!")


print("Checking contract date is always less or equal than application date")
cur.execute("""
SELECT
    COUNT(*)
FROM
    applications
JOIN
    contracts ON contracts.id_number = applications.id_number
WHERE
    contracts.contract_date < applications.application_date
""")
temp_count = cur.fetchone()[0]
if temp_count > 0:
    print("Found wrong contract_date (<application_date)!")
    cur.execute("DELETE FROM applications JOIN contracts ON contracts.id_number = applications.id_number WHERE contracts.contract_date < applications.application_date")
    print("Deleted applications with wrong date!")


print("-"*50)


print("Task3.\nConverting categorical data...")


applications_df = pd.read_sql_query("SELECT * FROM %s" % APPLICATIONS_TABLE, conn)


contracts_df = pd.read_sql_query("SELECT * FROM %s" % CONTRACTS_TABLE, conn)


for column_name in ['employed_by', 'education', 'marital_status', 'position', 'income_type', 'housing']:
    factorized_data, order = applications_df[column_name].factorize()
    factorized_data = factorized_data + 1
    factorized_data[factorized_data == 0] = factorized_data.max() + 1
    applications_df['new_%s' % column_name] = factorized_data
    cur.execute("ALTER TABLE %s DROP COLUMN %s" % (APPLICATIONS_TABLE, column_name))
    cur.execute("""CREATE TABLE %s_dict (
                        id INT AUTO_INCREMENT,
                        value VARCHAR(100),
                        PRIMARY KEY (id)
                    )  ENGINE=INNODB;""" % (column_name))
    cur.execute("""ALTER TABLE %s
                   ADD COLUMN %s_id INT""" % (APPLICATIONS_TABLE,
                                                               column_name))

    cur.execute("""ALTER TABLE %s
                   ADD CONSTRAINT fk_%s_id FOREIGN KEY (%s_id) REFERENCES %s(id)""" % (APPLICATIONS_TABLE,
                                                                                       column_name,
                                                                                       column_name,
                                                                                       column_name + "_dict"))

    for new_id, value in enumerate(order.tolist() + [None]):
        value = "'" + value + "'" if value else 'NULL'
        cur.execute("""INSERT INTO %s_dict VALUES (%d, %s)""" % (column_name, new_id + 1, value))

    for ind in applications_df.index:
        row = applications_df.loc[ind]
        cur.execute("UPDATE %s SET %s_id = %d WHERE id_number=%d" % (APPLICATIONS_TABLE,
                                                                     column_name,
                                                                     row["new_%s" % column_name],
                                                                     row.id_number))

for column_name in ['type']:
    factorized_data, order = contracts_df[column_name].factorize()
    factorized_data = factorized_data + 1
    factorized_data[factorized_data == 0] = factorized_data.max() + 1
    contracts_df['new_%s' % column_name] = factorized_data
    cur.execute("ALTER TABLE %s DROP COLUMN %s" % (CONTRACTS_TABLE, column_name))
    cur.execute("""CREATE TABLE %s_dict (
                        id INT AUTO_INCREMENT,
                        value VARCHAR(100),
                        PRIMARY KEY (id)
                    )  ENGINE=INNODB;""" % (column_name))
    cur.execute("""ALTER TABLE %s
                   ADD COLUMN %s_id INT""" % (CONTRACTS_TABLE,
                                                               column_name))

    cur.execute("""ALTER TABLE %s
                   ADD CONSTRAINT fk_%s_id FOREIGN KEY (%s_id) REFERENCES %s(id)""" % (CONTRACTS_TABLE,
                                                                                       column_name,
                                                                                       column_name,
                                                                                       column_name + "_dict"))

    for new_id, value in enumerate(order.tolist() + [None]):
        value = "'" + value + "'" if value else 'NULL'
        cur.execute("""INSERT INTO %s_dict VALUES (%d, %s)""" % (column_name, new_id + 1, value))

    for ind in contracts_df.index:
        row = contracts_df.loc[ind]
        cur.execute("UPDATE %s SET %s_id = %d WHERE id_number=%d" % (CONTRACTS_TABLE,
                                                                     column_name,
                                                                     row["new_%s" % column_name],
                                                                     row.id_number))
conn.commit()


print("Converted categorical data and created dictionary tables!")
